import csv
import re
import string
import xlwt
import datetime
import pandas as pd

from openpyxl import load_workbook
from openpyxl import Workbook

read_excel_file = '/Users/cheng/Desktop/Testing/Fall_2018_Results.xlsm'
write_excel_file = '/Users/cheng/Desktop/Testing/T.xlsx'

InputFile = read_excel_file
OutputFile = write_excel_file

# Get File Name based on comparison between exsiting titles and strings
def Get_Course_Name(InputFile, OutputFile, TitleFile = None):
    """
    Keyword Arguments:
        inputFile {str} -- path of file to parse
        OutputFile {str} -- path of file to save
        TitleFile {str} -- path of file with all titles (default: {None})
    """

    if TitleFile is None:
        TitleFile = '/Users/cheng/Desktop/Testing/Course_Title.xlsm'

    title = pd.DataFrame(pd.read_excel(TitleFile, keep_default_na=False))
    data = pd.DataFrame(pd.read_excel(InputFile, keep_default_na=False))
    wb = load_workbook(OutputFile)
    wb1 = wb.active
    Titles = []

    # get all titles:
    for i in range(title.iloc[:,0].size):
        buff = title.loc[i][0]
        buff = str(buff)
        Titles.append(buff)
        
    for i in range(len(data['file_beginning'])):
        text = ''
        length = 0
        info = data.loc[i][1]
        info = str(info)    
        
        for title in Titles:
            if title.lower() in info.lower():
                # useually the longest match string is the title for that course
                if len(title) > length:
                    text = title
                    length = len(title)
        
        wb1.cell(i+2,1,text)
      
    wb.save(write_excel_file)
    print('reday')


# get ID from file name ----------
def Get_ID(InputFile, OutputFile):
    """
    Keyword Arguments:
        inputFile {str} -- path of file to parse
        OutputFile {str} -- path of file to save
    """
    data = pd.DataFrame(pd.read_excel(InputFile, keep_default_na=False))
    wb = load_workbook(OutputFile)
    wb1 = wb.active

    for i in range(len(data['file_beginning'])):
        buff = data.loc[i][0]
        
        if buff:
            buff = str(buff)
            numbers = re.findall(r'\d+', buff)
            text = ''
            
            for number in numbers:
                num = int(number)
                if num >= 100 and num <= 9999 and num != 2015 and num != 2016 and num != 2017 and num != 2018 and num != 2019:
                    text = number
                    break

            wb1.cell(i+2,7,text)
            
    wb.save(write_excel_file)
    print('ready')

if __name__ == "__main__":
    Get_ID(InputFile, OutputFile)
    # Get_Course_Name(InputFile, OutputFile)


# the whole beginning ------- 0 ----------

# for i in range(3175):
#     buff = data.loc[i][2]
#     number = data.loc[i][1]
#     num = str(number)
#     text = ""
#     if buff:
#         buff = str(buff)
#         words = re.split('[\n]', buff)
#         final_words = []
        
#         for word in words:
#             if word.strip() != '':
#                 final_words.append(word)       
        
#         if num == '':
#             text = buff
#         elif num not in buff:
#             text = buff
#         else:
#             for j in range(len(final_words)):      length 60
#                 if num in final_words[j]:
#                     text = final_words[j]
#                     if j > 0:
#                         text = text + '\n' + final_words[j - 1]           add

#                     if j < len(final_words) - 1:
#                         text = text + '\n' + final_words[j + 1]
#                     break 
    
#     text = text.replace('\t', ' ')
#     wb1.cell(i+2,1,text)

# wb.save(write_excel_file)
# print('ready')

# remove parts for beginning ------ 1 ----------

# for i in range(3175):
#     text = data.loc[i][3]
    
#     if text:
#         text = str(text)
        
#     number = data.loc[i][1]
#     num = str(number)
    
#     if num in text:
#         text = re.sub(r'[A-Z]{2,7}\s+\d+\:?', "", text)        Section   Course Information and Policies (General Syllabus)     ECE 8930 007:    English 214    EDF9790          EdF 4800
# #     position = info.find(num)
    
#     if ':' in text:
#         text = re.sub(r'\d{1,2}\:\d{2}', "", text)
    
# #     if position <= 9:
# #         text = info[position + 4:] 
# #     else:
# #         text = info[:(position - 5)] + info[position + 4:]

#     if 'Dr.' in text:           
#         position0 = text.find('Dr.')
#         text = text[:position0]              Äò          -Fall     Syllabus for      

#     words = ['Credit Hour', 'credits', 'and RCID', 'GENERAL INFORMATION', 'General Information', 'Äê', 'Äù', 'Äî','Ä¢', 'Äú', 'àí', 'Äì', '¬≠', '†', '≠ ê', 'Æ', '¬', 'TR ', 'pm ', '[', ']', '_', 'Mon.', 'Wed.', 'Fri.', 'CI Team', 'TTh', 'Fall One', 'Fall Two', 'Mini A', 'ONLINE', 'Session', 'MINIMESTER A', 'Minimester C', 'Mini C', ';', '¬†', 'Minimester A', 'FALL II', 'FALL I', 'FALL', 'SYLLABUS', 'Full Term', 'Fall Minimester A', 'Course Syllabus', 'COURSE OUTLINE', 'Syllabus', 'Fall II','Fall I', 'Fall', '2015', 'CLEMSON UNIVERSITY', 'Clemson University', 'Section', 'section', 'Sec.', 'sec.']
#     for word in words:                                                                       
#         if word in text:                           
#             text = text.replace(word, "")                       Spring      
#         if 'Äô' in text:
#             text = text.replace('Äô', "‘")
#         if '¬†' in text:
#             text = text.replace('¬†', " ")
#              if '.' in text:
#                  text = text.replace('.', " ")        

#     wb1.cell(i+2,5,text)            
    
# wb.save(write_excel_file)
# print('ready')

# remove parts for beginning ------ 2 ----------

# for i in range(3175):
#     info = data.loc[i][3]
    
#     text = ''
#     if info:
#         text = str(info)
        
#     if '-' in text:
#         text = re.sub(r'\-\s?\d+', "", text)
        
#     if '/' in text:
#         text = re.sub(r'\/\d+', "", text)       

#     wb1.cell(i+2,7,text)            
    
# wb.save(write_excel_file)
# print('ready')


# remove rows ------ 3 ----------

# for i in range(3175):
#     buff = data.loc[i][3]
#     text = ""                                                                    Guidelines
#     if buff:                                             
#         buff = str(buff)
#         words = re.split('[\n]', buff)              SCHOOL COUNSELING
#         for word in words:
#             if len(word.strip()) > 4 and 'dr.' not in word and 'MSON UNIVERSITY' not in word and 'auditorium' not in word.lower() and 'about me' not in word.lower() and 'a.m.' not in word.lower() and 'p.m.' not in word.lower() and 'Tues.' not in word and 'Thur.' not in word and 'Tue ' not in word and 'Thu ' not in word and 'LEGE ' not in word and 'Mr.' not in word and 'Ms.' not in word and 'building' not in word.lower() and 'LEGE ' not in word and 'weekly' not in word.lower() and 'LEGE ' not in word and 'TBD' not in word and 'son University' not in word and 'manual' not in word.lower() and 'time' not in word.lower() and 'session' not in word.lower() and 'summer' not in word.lower() and 'credits' not in word.lower() and 'credit hr' not in word.lower() and 'credit units' not in word.lower() and 'schedule' not in word.lower() and 'dates' not in word.lower() and 'http' not in word.lower() and 'cell' not in word.lower() and 'lecture' not in word.lower() and 'meeting times' not in word.lower() and 'duration' not in word.lower() and 'advisor' not in word.lower() and 'coordinator' not in word.lower() and 'instructor' not in word.lower() and 'center' not in word.lower() and 'professor' not in word.lower() and 'course outline' not in word.lower() and 'requirements' not in word.lower() and 'course calendar' not in word.lower() and 'lecturer' not in word.lower() and 'prof ' not in word.lower() and 'm/w' not in word.lower() and 'teaching assistant' not in word.lower() and 't, th' not in word.lower() and 'mon/wed/fri' not in word.lower() and 't/th' not in word.lower() and 'm.w.f.' not in word.lower() and 'tu & th' not in word.lower() and 'tu ' not in word.lower() and 'm, w, f' not in word.lower() and 'wed.' not in word.lower() and 'mw:' not in word.lower() and 'semester' not in word.lower() and 'clemson university ' not in word.lower() and 'ph.d' not in word.lower() and 'prof.' not in word.lower() and 'room ' not in word.lower() and 'tth' not in word.lower() and 'hours' not in word.lower() and 'office' not in word.lower() and 'address' not in word.lower() and 'phone' not in word.lower() and 'hall' not in word.lower() and 'department' not in word.lower() and 'school' not in word.lower() and 'college' not in word.lower() and 'page' not in word.lower() and '@' not in word.lower() and 'august' not in word.lower() and 'mwf' not in word.lower() and 'monday' not in word.lower() and 'friday' not in word.lower() and 'tuesday' not in word.lower() and 'thursday' not in word.lower() and 'wednesday' not in word.lower() and 'location' not in word.lower() and 'tth ' not in word.lower() and 'contact ' not in word.lower():
#                 text = text + '\n' + word.strip()    Credit hour  credit
            
#             elif 'multiple sections' in word or 'Cancer Cell Comparisons' in word or 'General Chemistry 1st Semester' in word or 'Thesis Hours' in word or 'High School' in word or 'Public School' in word or 'Elementary School' in word or 'Primary School' in word or 'School Counseling' in word or 'School  Administrators' in word or 'School Administration' in word or 'PROFESSIONAL' in word or 'Introduction to the Course' in word or 'Middle School Curriculum' in word:
#                 text = text + '\n' + word.strip()
                                                                  
#             else:             
#                 continue                                       
# #         syllabus                 Course No.  Title     ê  ®   ë        TDA         'DR  '                             
#     wb1.cell(i+2,10,text)

# wb.save(write_excel_file)
# print('ready')


# remove parts for beginning ------ 4 ----------         ENGL 1030, Section 7:

# for i in range(3175):
#     text = data.loc[i][2]
    
#     if text:
#         text = str(text)
        
#     number = data.loc[i][0]
#     num = str(number)
    
#     if num in text:
#         text = re.sub(r'[A-Za-z]{2}\s?[A-Za-z]{2}\s{1}\d{4}', "", text)
        
#     if num in text:
#         text = re.sub(r'[A-Z]{2}\s?\d+\:?', "", text)
    
#     if '(' in text:
#         text = re.sub(r'\(\s?\d+\)', "", text)     

#     wb1.cell(i+2,22,text)            
    
# wb.save(write_excel_file)
# print('ready')


# remove rows with single word ------ 3 ----------

# for i in range(2610):
#     buff = data.loc[i][3]

#     text = ""
#     if buff:
#         buff = str(buff)
#         words = re.split('[\n\t]', buff)
#         for word in words:
#             word = word.strip()
#             if ' ' not in word and word.lower() != 'garth' and word.lower() != 'fenstermacher' and word.lower() != 'internship' and word.lower() != 'auditing' and word.lower() != 'psychology':
#                 continue
#             else:
#                 text = text + '\n' + word               
#                 wb1.cell(i+2,1,text)       Criminology     Drawing     DRAWING        Biofabrication     CUMUNC    Buddhism 

# wb.save(write_excel_file)
# print('ready')

# remove rows with "   " ------ 4 ----------

# for i in range(3175):
#     buff = data.loc[i][4]

#     text = ""
#     if buff:
#         buff = str(buff)
#         words = re.split('[\n]', buff)
#         for word in words:
#             word = word.strip()
#             if '   ' in word:           CREATIVE INQUIRY    DEVELOPMENT OF THE AVS CORNER      Independent Study    Graduate Level
#                 continue
#             else:
#                 text = text + '\n' + word               
#                 wb1.cell(i+2,7,text)

# wb.save(write_excel_file)
# print('ready')



# 'and GEN'


# remove parts for Course Title ------ 4 ----------

# for i in range(3175):
#     info = data.loc[i][3]
    
#     if info:
#         info = str(info)     Course Title/ Number	:       Course Title/Number	:
    
#     words = ['Course number and title', 'Number and Section', 'Section', 'Course title', 'Course Title and Number', 'Course Title', 'COURSE TITLE', 'Course Topic/Title', 'Course Title/Number', 'Course Title/ Number', ':']
     
#     for word in words:
#         if word in info:
#             info = info.replace(word, " ")
            
#     number = data.loc[i][1]
#     num = str(number)
# #     position = text.find(num)

#     if num in text:
#         text = re.sub(r'[A-Z]{2,7}\s+\d+\:?', "", text)
    
# #     if position <= 9:
# #         info = info[position + 4:] 
# #     else:
# #         info = info[:(position - 5)] + text[position + 4:]  
    
#     text = re.sub(u"\\(.*?\\)", "", info)
            
#     wb1.cell(i+2,16,text)         
    
# wb.save(write_excel_file)
# print('ready')

# Course Description ------ 5 ----------
# for i in range(232):
#     buff = data.loc[i][4]
#     number = data.loc[i][1]
#     num = str(number)

#     text = ""
#     if buff:
#         buff = str(buff)
#         words = re.split('[\n]', buff)
#         for word in words:
#             word = word.strip()
#             if num in word:
#                 text = text + '\n' + word
#                 wb1.cell(i+2,1,text)

# wb.save(write_excel_file)
# print('ready')

# syllabus  number

# Syllabus ------ 6 ----------
# for i in range(232):
#     buff = data.loc[i][4]
#     number = data.loc[i][1]
#     num = str(number)

#     text = ""
#     if buff:
#         buff = str(buff)
#         words = re.split('[\n]', buff)
#         for word in words:
#             word = word.strip()
#             if num in word:
#                 text = text + '\n' + word
#                 wb1.cell(i+2,1,text)

# wb.save(write_excel_file)
# print('ready')

# syllabus  number


# Syllabus ------ 7 ----------

# for i in range(3175):
#     buff = data.loc[i][7]

#     text = ""
#     if buff:
#         buff = str(buff)
#         words = re.split('[\n]', buff)
#         for word in words:
#             word = word.strip()
#             if 'Course	:' in word:
#                 text = text + '\n' + word
#                 break
                
#             if 'Title	:' in word:
#                 text = text + '\n' + word            
#                 break
 
#      Syllabus	:       Topic	:
     

#     wb1.cell(i+2,20,text)

# wb.save(write_excel_file)             Course	:
# print('ready')
